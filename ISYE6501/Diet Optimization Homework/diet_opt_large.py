from pulp import *
from openpyxl import load_workbook

# open and load diet excel file to read date
book = load_workbook('diet_large.xlsx')
sheet = book['Sheet1']

# define empty lists to fill with excel data to use as variables in optimization model
nutr = []
foods = []
price_per_serving = []
cholesterol = []
nutr_min = []
nutr_max = []

# create lists of data from excel

# nutrient list
for col in sheet.columns:
    nutr.append(col[1].value)
del nutr[0]
#print(nutr)

# list of nutr min
for col in sheet.columns:
    nutr_min.append(col[7149].value)
del nutr_min[0]
del nutr_min[-3:]
#print(nutr_min)

# list of nutr max
for col in sheet.columns:
    nutr_max.append(col[7151].value)
del nutr_max[0]
del nutr_max[-3:]
#print(nutr_max)

# food list
for row in sheet.rows:
    foods.append(row[0].value)
del foods[:2]
del foods[-4:]
#print(foods)

# cholesterol values list <--- minimization problem
for row in sheet.rows:
    cholesterol.append(row[28].value)
del cholesterol[:2]
del cholesterol[-4:]
cholesterol = [0 if v is None else v for v in cholesterol]
#print(cholesterol)

# dictionary of food with cholesterol values
cholesterol_per_food_dict = {foods[i]:cholesterol[i] for i in range(len(foods))}

# dictionary of food with all nutrional values
food_nutr_dict = {foods[i] : {nutr[j]:sheet.cell(row = i+3, column = j + 2).value for j in range(len(nutr))} for i in range(len(foods))}

# Setting up the problem <-- Mihnimizing the cholesterol of the diet
prob = LpProblem('MinimumCholesterol', LpMinimize)

# Setting up the our decision variables(s)
food_vars = LpVariable.dicts("AmountFood", foods, 0)

# Setting up objective function
prob += lpSum(cholesterol_per_food_dict[f] * food_vars[f] for f in foods)

# Constraints min/max value for each nutrient
for i in range(0,27):
    min_check = lpSum(food_nutr_dict[f][nutr[i]] * food_vars[f] for f in foods)
    min_condition = nutr_min[i] <= min_check
    prob += min_condition

for i in range(0,27):
    max_check = lpSum(food_nutr_dict[f][nutr[i]] * food_vars[f] for f in foods)
    max_condition = nutr_max[i] >= max_check
    prob += max_condition

# Solve
prob.solve()
print("Status", LpStatus[prob.status])

for v in prob.variables():
    if v.varValue > 0:
        print(v.name, '=', v.varValue)
    else:
        pass

print("Total cholesterol of diet =", value(prob.objective))
