from pulp import *
from openpyxl import load_workbook

# open and load diet excel file to read date
book = load_workbook('diet.xlsx')
sheet = book['Sheet1']

#define empty lists to fill with excel data to use as variables in optimization model
nutr = []
foods = []
price_per_serving = []

# list of nutrients
for col in sheet.columns:
    nutr.append(col[0].value)
del nutr[0:3]

# list of foods
for row in sheet.rows:
    #print (row[0].value)
    foods.append(row[0].value)
del foods[0]
del foods[-3:]

# list of price per serving
for row in sheet.rows:
    price_per_serving.append(row[1].value)
del price_per_serving[0]
del price_per_serving[-3:]

# dictionary of serving size for each food
serving_size = {foods[i]: sheet.cell(row= i + 2, column = 3).value for i in range(len(foods))}

# dictionary of price per serving of food
price_per_serving_dict = {foods[i]:price_per_serving[i] for i in range(len(foods))}

# create dictionary for all nutrients
#['Calories', 'Cholesterol mg', 'Total_Fat g', 'Sodium mg', 'Carbohydrates g', 'Dietary_Fiber g', 'Protein g', 'Vit_A IU', 'Vit_C IU', 'Calcium mg', 'Iron mg']

# dictionary of calories per serving of food
calories = {foods[i]: sheet.cell(row = i + 2, column = 4).value for i in range(len(foods))}

# dictionary of cholesterol
cholesterol = {foods[i]: sheet.cell(row = i +2, column = 5).value for i in range(len(foods))}

# dictionary of total fat
total_fat = {foods[i]: sheet.cell(row = i +2, column = 6).value for i in range(len(foods))}

# dictionary for sodium
sodium = {foods[i]: sheet.cell(row = i +2, column = 7).value for i in range(len(foods))}

# dictionary for carbs
carbs = {foods[i]: sheet.cell(row = i +2, column = 8).value for i in range(len(foods))}

# dictionary for fiber
fiber = {foods[i]: sheet.cell(row = i +2, column = 9).value for i in range(len(foods))}

# dictionary for protein
protein = {foods[i]: sheet.cell(row = i +2, column = 10).value for i in range(len(foods))}

# dictionary for vit_A
vit_A = {foods[i]: sheet.cell(row = i +2, column = 11).value for i in range(len(foods))}

# dictionary for vit_C
vit_C = {foods[i]: sheet.cell(row = i +2, column = 12).value for i in range(len(foods))}

# dictionary for calcium
calcium = {foods[i]: sheet.cell(row = i +2, column = 13).value for i in range(len(foods))}

# dictionary for iron
iron = {foods[i]: sheet.cell(row = i +2, column = 14).value for i in range(len(foods))}


# dictionary of food with its nutrient values
# food_nutr_dict = {foods[i] : { nutr[j]: sheet.cell(row = i + 2, column = j + 4).value for j in range(len(nutr))} for i in range(len(foods))}

# Setting up the problem <-- Minimizing the cost of the diet
prob = LpProblem('Diet', LpMinimize)


# Setting up our decision variables
#1. we do not have to select all of the foods, so we will set a decision variable for either using the food or not using the food (binary decision)
chosen_food_vars = LpVariable.dicts("Food Chosen", foods, 0, 1, cat = "Integer")

#2. choosing foods variable decision
food_vars = LpVariable.dicts("Food", foods, 0)

# if a food is chosen at least 1/10 of the serving size must be selected
for f in foods:
    prob += food_vars[f] >= chosen_food_vars[f]*0.1
    prob += food_vars[f] <= chosen_food_vars[f]*1000000


# Setting up objective function
prob += lpSum(price_per_serving_dict[f] * food_vars[f] for f in foods)

#Constraints
# min/max of calories
prob += lpSum(calories[i]*food_vars[i] for i in foods) >= 1500
prob += lpSum(calories[i]*food_vars[i] for i in foods) <= 2500

# min/max of cholesterol
prob += lpSum(cholesterol[i]*food_vars[i] for i in foods) >= 30
prob += lpSum(cholesterol[i]*food_vars[i] for i in foods) <= 240

# min/max of total_fat
prob += lpSum(total_fat[i]*food_vars[i] for i in foods) >= 20
prob += lpSum(total_fat[i]*food_vars[i] for i in foods) <= 70

# min/max of sodium
prob += lpSum(sodium[i]*food_vars[i] for i in foods) >= 800
prob += lpSum(sodium[i]*food_vars[i] for i in foods) <= 2000

# min/max of carbs
prob += lpSum(carbs[i]*food_vars[i] for i in foods) >= 130
prob += lpSum(carbs[i]*food_vars[i] for i in foods) <= 450

# min/max of fiber
prob += lpSum(fiber[i]*food_vars[i] for i in foods) >= 125
prob += lpSum(fiber[i]*food_vars[i] for i in foods) <= 250

# min/max of protein
prob += lpSum(protein[i]*food_vars[i] for i in foods) >= 60
prob += lpSum(protein[i]*food_vars[i] for i in foods) <= 100

# min/max of vit_A
prob += lpSum(vit_A[i]*food_vars[i] for i in foods) >= 1000
prob += lpSum(vit_A[i]*food_vars[i] for i in foods) <= 10000

# min/max of vit_C
prob += lpSum(vit_C[i]*food_vars[i] for i in foods) >= 400
prob += lpSum(vit_C[i]*food_vars[i] for i in foods) <= 5000

# min/max of calcium
prob += lpSum(calcium[i]*food_vars[i] for i in foods) >= 700
prob += lpSum(calcium[i]*food_vars[i] for i in foods) <= 1500

# min/max of iron
prob += lpSum(iron[i]*food_vars[i] for i in foods) >= 10
prob += lpSum(iron[i]*food_vars[i] for i in foods) <= 40

# choose only celery or broccoli
prob += chosen_food_vars['Frozen Broccoli'] + chosen_food_vars['Celery, Raw'] <= 1

# at least 3 kinds of meat/poultry/fish/eggs
prob += chosen_food_vars['Roasted Chicken'] + chosen_food_vars['Poached Eggs'] + chosen_food_vars['Scrambled Eggs'] + chosen_food_vars['Bologna,Turkey'] + chosen_food_vars['Frankfurter, Beef'] + chosen_food_vars['Ham,Sliced,Extralean'] + chosen_food_vars['Kielbasa,Prk'] + chosen_food_vars['Pizza W/Pepperoni'] + chosen_food_vars['Taco'] + chosen_food_vars['Hamburger W/Toppings'] + chosen_food_vars['Hotdog, Plain'] + chosen_food_vars['Pork'] + chosen_food_vars['Sardines in Oil'] + chosen_food_vars['White Tuna in Water'] + chosen_food_vars['Chicknoodl Soup'] + chosen_food_vars['Splt Pea&Hamsoup'] + chosen_food_vars['Vegetbeef Soup'] + chosen_food_vars['Beanbacn Soup,W/Watr'] >= 3

# Solve
prob.solve()
print("Status", LpStatus[prob.status])

for v in prob.variables():
    if v.varValue > 0:
        print(v.name, '=', v.varValue)
    else:
        pass

print("Total cost of diet =", value(prob.objective))
