from pulp import *
from openpyxl import load_workbook

# open and load diet excel file to read date
book = load_workbook('diet.xlsx')
sheet = book['Sheet1']

#define empty lists to fill with excel data to use as variables in optimization model
nutr = []
foods = []
price_per_serving = []
nutr_min = []
nutr_max = []

# list of nutrients
for col in sheet.columns:
    nutr.append(col[0].value)
del nutr[0:3]

# list of nutrient min
for i in range(0,11):
    nutr_min.append(sheet.cell(row = 67, column = i + 4).value)
print(nutr_min)

# list of nutrient max
for i in range(0,11):
    nutr_max.append(sheet.cell(row = 68, column = i +4).value)
print(nutr_max)

# list of foods
for row in sheet.rows:
    #print (row[0].value)
    foods.append(row[0].value)
del foods[0]
del foods[-3:]

# list of price per serving
for row in sheet.rows:
    price_per_serving.append(row[1].value)
del price_per_serving[0]
del price_per_serving[-3:]

# dictionary of price per serving of food
price_per_serving_dict = {foods[i]:price_per_serving[i] for i in range(len(foods))}

# dictionary of food with its nutrient values
food_nutr_dict = {foods[i] : { nutr[j]: sheet.cell(row = i + 2, column = j + 4).value for j in range(len(nutr))} for i in range(len(foods))}

# Setting up the problem <-- Minimizing the cost of the diet
prob = LpProblem('Diet', LpMinimize)


# Setting up our decision variables
food_vars = LpVariable.dicts("AmountFood", foods, 0)

# Setting up objective function
prob += lpSum(price_per_serving_dict[f] * food_vars[f] for f in foods)

#Constraints min/max value for each nutrient
for i in range(0,11):
    min_check = lpSum(food_nutr_dict[f][nutr[i]] * food_vars[f] for f in foods)
    min_condition = nutr_min[i] <= min_check
    prob += min_condition

for i in range(0,11):
    max_check = lpSum(food_nutr_dict[f][nutr[i]] * food_vars[f] for f in foods)
    max_condition = nutr_max[i] >= max_check
    prob += max_condition

# Solve
prob.solve()
print("Status", LpStatus[prob.status])

for v in prob.variables():
    if v.varValue > 0:
        print(v.name, '=', v.varValue)
    else:
        pass

print("Total cost of diet =", value(prob.objective))
